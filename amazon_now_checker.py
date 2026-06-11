"""
Amazon Now / Fresh Pincode Serviceability Checker
===================================================
Reads pincodes from 'quick commerce dark store.xlsx', checks each against
Amazon Fresh (formerly Amazon Now) using Playwright browser automation,
and outputs results to 'amazon_now_results.xlsx'.

Approach:
  1. Navigate to Amazon Fresh homepage.
  2. For each pincode, use the location-change popup to set the delivery pincode.
  3. After setting, check if Amazon Fresh content loads (serviceable) or if
     Amazon shows "not available in your area" messaging.
  4. Record result with timestamp.

Anti-bot measures:
  - Realistic random delays between requests (3-8 seconds).
  - Human-like mouse movements and typing speed.
  - Session persistence via storage_state.
  - Incremental saves every 25 pincodes.
  - Automatic retry on failures with exponential backoff.

Usage:
  Step 1 (first time only): Run with --login flag to save session
    python amazon_now_checker.py --login

  Step 2: Run the checker
    python amazon_now_checker.py

  Resume after interruption (skips already-checked pincodes):
    python amazon_now_checker.py --resume
"""

import argparse
import json
import os
import random
import sys
import time
from datetime import datetime, timezone, timedelta

import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeout

# --- Configuration -----------------------------------------------------------
INPUT_FILE = "quick commerce dark store.xlsx"
OUTPUT_FILE = "amazon_now_results.xlsx"
PROGRESS_FILE = "progress.json"
SESSION_FILE = "amazon_session.json"

# Amazon Fresh URL (integrated into main Amazon.in site)
AMAZON_FRESH_URL = "https://www.amazon.in/tez/browse/home?qcbrand=qqfsWw9RkO&ref=nav_cs_dsk_grfl_stfr_at"
AMAZON_HOME_URL = "https://www.amazon.in"

# Delays (seconds) -- randomised within these ranges
MIN_DELAY = 1
MAX_DELAY = 3
TYPING_DELAY_MS = 40  # ms between keystrokes (human-like)


# Save progress every N pincodes
SAVE_INTERVAL = 25

# Max retries per pincode
MAX_RETRIES = 3

# IST timezone
IST = timezone(timedelta(hours=5, minutes=30))


# --- Helpers -----------------------------------------------------------------
def human_delay(low=MIN_DELAY, high=MAX_DELAY):
    """Sleep for a random human-like duration."""
    time.sleep(random.uniform(low, high))


def short_delay():
    """Brief delay for UI transitions."""
    time.sleep(random.uniform(0.5, 1.0))


def load_pincodes(filepath: str) -> list[dict]:
    """
    Load all rows from every sheet in the input Excel.
    Returns list of dicts with keys: sheet, city, locality, pincode.
    """
    wb = openpyxl.load_workbook(filepath, read_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            if row[2] is not None:
                rows.append({
                    "sheet": sheet_name.strip(),
                    "city": str(row[0]).strip() if row[0] else "",
                    "locality": str(row[1]).strip() if row[1] else "",
                    "pincode": str(int(row[2])),
                })
    wb.close()
    return rows


def get_unique_pincodes(rows: list[dict]) -> list[str]:
    """Return deduplicated list of pincodes preserving first-seen order."""
    seen = set()
    unique = []
    for r in rows:
        pc = r["pincode"]
        if pc not in seen:
            seen.add(pc)
            unique.append(pc)
    return unique


def load_progress() -> dict:
    """Load progress from JSON file."""
    if os.path.exists(PROGRESS_FILE):
        with open(PROGRESS_FILE, "r") as f:
            return json.load(f)
    return {}


def save_progress(progress: dict):
    """Save progress to JSON file."""
    with open(PROGRESS_FILE, "w") as f:
        json.dump(progress, f, indent=2)


def save_results_excel(results: list[dict], all_rows: list[dict]):
    """
    Save results to Excel with columns:
      pincode | city | locality | source_platform | amazon_now_serviceable | checked_at
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Amazon Now Results"

    # Headers
    headers = [
        "Pincode",
        "City",
        "Locality",
        "Source Platform",
        "Amazon Now Serviceable",
        "Checked At (IST)",
    ]
    ws.append(headers)

    # Style headers
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Build a lookup from pincode -> result
    result_lookup = {}
    for r in results:
        result_lookup[r["pincode"]] = r

    # Write data rows -- one row per original row in the input
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    row_num = 2
    for entry in all_rows:
        pc = entry["pincode"]
        res = result_lookup.get(pc, {})
        serviceable = res.get("serviceable", "Not Checked")
        checked_at = res.get("checked_at", "")

        # Map sheet name to platform
        sheet = entry["sheet"].lower()
        if "blinkit" in sheet:
            platform = "Blinkit"
        elif "swiggy" in sheet:
            platform = "Swiggy Instamart"
        elif "zepto" in sheet:
            platform = "Zepto"
        else:
            platform = entry["sheet"]

        ws.cell(row=row_num, column=1, value=pc).border = thin_border
        ws.cell(row=row_num, column=2, value=entry["city"]).border = thin_border
        ws.cell(row=row_num, column=3, value=entry["locality"]).border = thin_border
        ws.cell(row=row_num, column=4, value=platform).border = thin_border

        svc_cell = ws.cell(row=row_num, column=5, value=serviceable)
        svc_cell.border = thin_border
        svc_cell.alignment = Alignment(horizontal="center")
        if serviceable is True:
            svc_cell.value = "TRUE"
            svc_cell.fill = green_fill
        elif serviceable is False:
            svc_cell.value = "FALSE"
            svc_cell.fill = red_fill
        else:
            svc_cell.fill = yellow_fill

        ws.cell(row=row_num, column=6, value=checked_at).border = thin_border
        row_num += 1

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 24

    # -- Summary sheet --------------------------------------------------------
    ws2 = wb.create_sheet("Summary")
    ws2.append(["Metric", "Value"])
    ws2.cell(row=1, column=1).font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2).font = header_font
    ws2.cell(row=1, column=2).fill = header_fill

    total_unique = len(set(r["pincode"] for r in all_rows))
    total_checked = len([r for r in results if r.get("serviceable") is not None])
    total_true = len([r for r in results if r.get("serviceable") is True])
    total_false = len([r for r in results if r.get("serviceable") is False])
    total_error = len([r for r in results if r.get("serviceable") is None])

    summary_data = [
        ("Total Unique Pincodes", total_unique),
        ("Pincodes Checked", total_checked),
        ("Amazon Now Serviceable (TRUE)", total_true),
        ("Amazon Now NOT Serviceable (FALSE)", total_false),
        ("Errors / Not Checked", total_error),
        ("Check Completed At", datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S IST")),
    ]
    for i, (metric, value) in enumerate(summary_data, 2):
        ws2.cell(row=i, column=1, value=metric)
        ws2.cell(row=i, column=2, value=value)

    ws2.column_dimensions["A"].width = 40
    ws2.column_dimensions["B"].width = 30

    # -- Unique Pincodes sheet ------------------------------------------------
    ws3 = wb.create_sheet("Unique Pincodes")
    ws3.append(["Pincode", "Amazon Now Serviceable", "Checked At (IST)"])
    for col_num in range(1, 4):
        cell = ws3.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for i, res in enumerate(results, 2):
        ws3.cell(row=i, column=1, value=res["pincode"])
        svc_cell = ws3.cell(row=i, column=2)
        if res.get("serviceable") is True:
            svc_cell.value = "TRUE"
            svc_cell.fill = green_fill
        elif res.get("serviceable") is False:
            svc_cell.value = "FALSE"
            svc_cell.fill = red_fill
        else:
            svc_cell.value = "ERROR"
            svc_cell.fill = yellow_fill
        svc_cell.alignment = Alignment(horizontal="center")
        ws3.cell(row=i, column=3, value=res.get("checked_at", ""))

    ws3.column_dimensions["A"].width = 12
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 24

    try:
        wb.save(OUTPUT_FILE)
        print(f"  [OK] Results saved to {OUTPUT_FILE}")
    except PermissionError:
        # File is likely open in Excel — save to a backup instead
        backup = OUTPUT_FILE.replace(".xlsx", "_backup.xlsx")
        wb.save(backup)
        print(f"  [WARN] {OUTPUT_FILE} is locked (open in Excel?). Saved to {backup} instead.")
        print(f"         Close Excel and the next save will use the original file.")


# --- Core Checker ------------------------------------------------------------
class AmazonFreshChecker:
    """Uses Playwright to check Amazon Fresh serviceability per pincode."""

    def __init__(self, headless: bool = False):
        self.headless = headless
        self.playwright = None
        self.browser = None
        self.context = None
        self.page = None

    def start(self, use_session: bool = True):
        """Launch browser and set up context."""
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(
            headless=self.headless,
            args=[
                "--disable-blink-features=AutomationControlled",
                "--no-sandbox",
            ],
        )

        context_opts = {
            "viewport": {"width": 1366, "height": 768},
            "user_agent": (
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/131.0.0.0 Safari/537.36"
            ),
            "locale": "en-IN",
            "timezone_id": "Asia/Kolkata",
        }

        if use_session and os.path.exists(SESSION_FILE):
            context_opts["storage_state"] = SESSION_FILE
            print(f"  [SESSION] Loaded saved session from {SESSION_FILE}")

        self.context = self.browser.new_context(**context_opts)
        self.page = self.context.new_page()

        # Block unnecessary resources to speed things up
        self.page.route(
            "**/*.{png,jpg,jpeg,gif,webp,svg,ico,woff,woff2,ttf,eot}",
            lambda route: route.abort(),
        )

    def stop(self):
        """Clean up browser resources."""
        if self.context:
            try:
                self.context.storage_state(path=SESSION_FILE)
                print(f"  [SAVE] Session saved to {SESSION_FILE}")
            except Exception:
                pass
        if self.browser:
            self.browser.close()
        if self.playwright:
            self.playwright.stop()

    def login_flow(self):
        """
        Interactive login: opens Amazon, waits for user to log in manually,
        then saves the session.
        """
        print("\n[LOGIN MODE]")
        print("=" * 60)
        self.start(use_session=False)

        # Unblock images for login
        self.page.unroute("**/*.{png,jpg,jpeg,gif,webp,svg,ico,woff,woff2,ttf,eot}")

        self.page.goto(AMAZON_HOME_URL, wait_until="domcontentloaded")
        print("  Amazon homepage loaded.")
        print("  --> Please log in manually in the browser window.")
        print("  --> After logging in, navigate to Amazon Fresh if you want.")
        print("  --> Then come back here and press ENTER to save the session.\n")

        input("  Press ENTER after logging in... ")

        self.context.storage_state(path=SESSION_FILE)
        print(f"\n  [OK] Session saved to {SESSION_FILE}")
        self.stop()

    def navigate_to_fresh(self):
        """Navigate to Amazon Home to set the pincode."""
        print("  [NAV] Navigating to Amazon Home...")
        self.page.goto(AMAZON_HOME_URL, wait_until="domcontentloaded", timeout=30000)
        short_delay()

    def set_pincode(self, pincode: str) -> bool:
        """
        Set the delivery pincode via Amazon's location popup.
        Returns True if the pincode was successfully set.
        """
        try:
            # Click on the "Deliver to" / location link
            location_link = self.page.locator("#nav-global-location-popover-link")
            if location_link.is_visible(timeout=5000):
                location_link.click()
                short_delay()
            else:
                # Try alternative selector
                alt_link = self.page.locator("#glow-ingress-line1")
                if alt_link.is_visible(timeout=3000):
                    alt_link.click()
                    short_delay()
                else:
                    print(f"    [WARN] Could not find location selector")
                    return False

            # Wait for the pincode input modal
            pincode_input = self.page.locator("#GLUXZipUpdateInput")
            pincode_input.wait_for(state="visible", timeout=5000)

            # Clear existing value and type new pincode
            pincode_input.fill("")
            short_delay()
            pincode_input.type(pincode, delay=TYPING_DELAY_MS)
            short_delay()

            # Click Apply
            apply_btn = self.page.locator(
                "input[aria-labelledby='GLUXZipUpdate-announce'],"
                "#GLUXZipUpdate span input,"
                "#GLUXZipUpdate"
            ).first
            apply_btn.click()
            time.sleep(1)

            # Sometimes a "Done" or "Continue" button appears after applying
            try:
                done_btn = self.page.locator(
                    "button[name='glowDoneButton'],"
                    "#GLUXConfirmClose,"
                    ".a-popover-footer button"
                ).first
                if done_btn.is_visible(timeout=3000):
                    done_btn.click()
                    short_delay()
            except PlaywrightTimeout:
                pass

            # Wait for page to update
            time.sleep(2)
            
            # Verify the pincode was actually set
            try:
                loc_text = self.page.locator("#glow-ingress-line2").inner_text(timeout=3000)
                if pincode not in loc_text.replace(" ", ""):
                    print(f"    [WARN] Pincode {pincode} not found in location text: {loc_text}")
                    return False
            except Exception:
                pass # If we can't find the element, just proceed and hope for the best
                
            return True

        except PlaywrightTimeout as e:
            print(f"    [TIMEOUT] Setting pincode {pincode}: {e}")
            return False
        except Exception as e:
            print(f"    [ERROR] Setting pincode {pincode}: {e}")
            return False

    def check_fresh_serviceable(self, pincode: str) -> bool | None:
        """
        After setting the pincode, check if Amazon Fresh is serviceable.

        Strategy:
          1. Navigate to the Fresh storefront with the pincode set.
          2. Look for indicators:
             - Serviceable: product listings, category grids, "Fresh" branding.
             - Not serviceable: "not available", "doesn't deliver", error banners.

        Returns:
          True  = Amazon Fresh is available for this pincode
          False = Amazon Fresh is NOT available
          None  = Could not determine (error)
        """
        try:
            # Navigate to Fresh page (which will respect the set pincode)
            self.page.goto(AMAZON_FRESH_URL, wait_until="domcontentloaded", timeout=30000)
            time.sleep(2)

            page_content = self.page.content().lower()

            # -- Positive signals (Fresh is available) ------------------------
            positive_signals = [
                "add to cart",
                "fresh-storefront",
                "freshstorefront",
                "grocery",
                "fruits",
                "vegetables",
                "dairy",
                "fresh-shoveler",
                "asin",  # product IDs visible = products are loading
            ]

            # -- Negative signals (Fresh NOT available) -----------------------
            negative_signals = [
                "not available in your area",
                "doesn't deliver to this",
                "does not deliver",
                "not available at this",
                "currently unavailable in",
                "is not available",
                "service is not available",
                "we don't deliver",
                "pincode is not serviceable",
                "not serviceable",
                "select a different location",
                "change your delivery address",
                "sorry, this service isn't available",
                "amazon fresh is not available",
                "amazon fresh isn't available",
                "this service is currently not available",
            ]

            # Count signals
            pos_count = sum(1 for s in positive_signals if s in page_content)
            neg_count = sum(1 for s in negative_signals if s in page_content)

            # Check for product cards (strong positive signal)
            product_cards = self.page.locator(
                "[data-asin]:not([data-asin=''])"
            ).count()

            # Check for price elements (strong positive signal)
            price_elements = self.page.locator(".a-price, .a-price-whole").count()

            # Strong negative signals — these specifically say Fresh is unavailable
            strong_negatives = [
                "amazon fresh is not available",
                "amazon fresh isn't available",
                "sorry, this service isn't available",
                "this service is currently not available",
                "we don't deliver",
                "not available in your area",
                "pincode is not serviceable",
            ]
            strong_neg_count = sum(1 for s in strong_negatives if s in page_content)

            # Decision logic:
            # 1. Product cards + prices = strongest signal that Fresh is live
            if product_cards >= 3 and price_elements >= 2:
                return True

            # 2. Strong negatives explicitly say Fresh is not available
            if strong_neg_count > 0 and product_cards < 2:
                return False

            # 3. Multiple positive text signals
            if pos_count >= 3 and product_cards >= 1:
                return True

            # 4. Has "add to cart" = products are purchasable
            if "add to cart" in page_content and product_cards >= 1:
                return True

            # 5. Weak negatives (like "is not available" for single items)
            #    only matter if there are no products at all
            if product_cards == 0 and price_elements == 0:
                if neg_count > 0:
                    return False
                return False  # No products, no prices = not serviceable

            # 6. Has some products/prices = likely serviceable
            if product_cards >= 1 or price_elements >= 2:
                return True

            return False

        except PlaywrightTimeout:
            print(f"    [TIMEOUT] Page timeout for pincode {pincode}")
            return None
        except Exception as e:
            print(f"    [ERROR] Checking pincode {pincode}: {e}")
            return None

    def check_pincode(self, pincode: str) -> dict:
        """
        Full flow: set pincode -> check serviceability -> return result.
        """
        result = {
            "pincode": pincode,
            "serviceable": None,
            "checked_at": datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S"),
        }

        for attempt in range(1, MAX_RETRIES + 1):
            try:
                # Set pincode
                if not self.set_pincode(pincode):
                    if attempt < MAX_RETRIES:
                        print(f"    [RETRY] {attempt}/{MAX_RETRIES} for {pincode}...")
                        human_delay(5, 10)
                        self.navigate_to_fresh()
                        continue
                    else:
                        return result

                # Check serviceability
                serviceable = self.check_fresh_serviceable(pincode)

                if serviceable is not None:
                    result["serviceable"] = serviceable
                    result["checked_at"] = datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")
                    return result

                if attempt < MAX_RETRIES:
                    print(f"    [RETRY] {attempt}/{MAX_RETRIES} for {pincode}...")
                    human_delay(5, 10)
                    self.navigate_to_fresh()

            except Exception as e:
                print(f"    [FAIL] Attempt {attempt} error for {pincode}: {e}")
                if attempt < MAX_RETRIES:
                    human_delay(5, 15)
                    try:
                        self.navigate_to_fresh()
                    except Exception:
                        pass

        return result


# --- Main --------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Check Amazon Fresh serviceability for pincodes"
    )
    parser.add_argument(
        "--login",
        action="store_true",
        help="Run in login mode -- opens browser for manual Amazon login",
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="Resume from last checkpoint (skip already-checked pincodes)",
    )
    parser.add_argument(
        "--headless",
        action="store_true",
        help="Run browser in headless mode (not recommended -- may trigger bot detection)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Limit the number of pincodes to check (0 = all)",
    )
    args = parser.parse_args()

    # -- Login mode -----------------------------------------------------------
    if args.login:
        checker = AmazonFreshChecker(headless=False)
        checker.login_flow()
        return

    # -- Load input data ------------------------------------------------------
    print("\n" + "=" * 60)
    print("  Amazon Now / Fresh -- Pincode Serviceability Checker")
    print("=" * 60)

    if not os.path.exists(INPUT_FILE):
        print(f"\n  [ERROR] Input file not found: {INPUT_FILE}")
        print(f"     Place '{INPUT_FILE}' in the current directory.")
        sys.exit(1)

    all_rows = load_pincodes(INPUT_FILE)
    unique_pincodes = get_unique_pincodes(all_rows)
    print(f"\n  [INFO] Input file: {INPUT_FILE}")
    print(f"     Total rows: {len(all_rows)}")
    print(f"     Unique pincodes: {len(unique_pincodes)}")

    # -- Resume logic ---------------------------------------------------------
    progress = load_progress() if args.resume else {}
    results = []

    if progress:
        # Rebuild results from progress
        for pc, data in progress.items():
            results.append(data)
        print(f"     Already checked: {len(results)} pincodes (resuming)")

    already_checked = set(progress.keys())
    to_check = [pc for pc in unique_pincodes if pc not in already_checked]

    if args.limit > 0:
        to_check = to_check[: args.limit]

    print(f"     Remaining to check: {len(to_check)}")
    print()

    if not to_check:
        print("  [OK] All pincodes have been checked!")
        save_results_excel(results, all_rows)
        return

    if not os.path.exists(SESSION_FILE):
        print("  [WARN] No saved session found. Run with --login first:")
        print("       python amazon_now_checker.py --login")
        print()
        response = input("  Continue without login? (y/N): ").strip().lower()
        if response != "y":
            return

    # -- Run checker ----------------------------------------------------------
    checker = AmazonFreshChecker(headless=args.headless)
    try:
        checker.start(use_session=True)
        checker.navigate_to_fresh()
        human_delay(2, 4)

        start_time = time.time()

        for i, pincode in enumerate(to_check):
            elapsed = time.time() - start_time
            rate = (i / elapsed * 3600) if elapsed > 0 and i > 0 else 0
            eta_mins = ((len(to_check) - i) / (i / elapsed / 60)) if i > 0 and elapsed > 0 else 0

            print(
                f"  [{i+1}/{len(to_check)}] Checking pincode {pincode}  "
                f"({rate:.0f}/hr, ETA: {eta_mins:.0f} min)"
            )

            result = checker.check_pincode(pincode)
            results.append(result)
            progress[pincode] = result

            if result["serviceable"] is True:
                status = "[YES]"
            elif result["serviceable"] is False:
                status = "[NO]"
            else:
                status = "[ERROR]"
            print(f"         -> Amazon Fresh: {status}")

            # Save progress periodically
            if (i + 1) % SAVE_INTERVAL == 0:
                save_progress(progress)
                save_results_excel(results, all_rows)
                print(f"  [SAVE] Progress saved ({i+1} pincodes checked)")

            # Human-like delay between checks
            if i < len(to_check) - 1:
                human_delay()

    except KeyboardInterrupt:
        print("\n\n  [PAUSED] Interrupted! Saving progress...")
    except Exception as e:
        print(f"\n  [FATAL] Error: {e}")
        print("     Saving progress...")
    finally:
        # Always save on exit
        save_progress(progress)
        save_results_excel(results, all_rows)
        checker.stop()

    # -- Final summary --------------------------------------------------------
    total_checked = len([r for r in results if r.get("serviceable") is not None])
    total_yes = len([r for r in results if r.get("serviceable") is True])
    total_no = len([r for r in results if r.get("serviceable") is False])
    total_err = len([r for r in results if r.get("serviceable") is None])

    print("\n" + "=" * 60)
    print("  RESULTS SUMMARY")
    print("=" * 60)
    print(f"  Total checked:       {total_checked}")
    print(f"  Serviceable (YES):   {total_yes}")
    print(f"  Not serviceable:     {total_no}")
    print(f"  Errors:              {total_err}")
    print(f"  Output file:         {OUTPUT_FILE}")
    print("=" * 60 + "\n")


if __name__ == "__main__":
    main()
